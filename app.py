"""
=============================================================================
OceENS - Application principale FastAPI (version fusionnée)
=============================================================================
Combine :
- L'authentification Azure Entra ID (auth.py)
- La gestion des sessions (SessionMiddleware)
- Les routes du module app (1).py (surveys, questionnaires, API)
- Les dashboards par rôle
"""

from dotenv import load_dotenv


import os
import io
from typing import Annotated, Dict, List, Optional
from datetime import datetime
from contextlib import asynccontextmanager

from fastapi import Depends, FastAPI, File, Form, Request, UploadFile, APIRouter
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, SQLModel, create_engine, select, func
import uvicorn
from seed import seed_all_if_necessary

# ┌─ Importation des modèles et du module d'authentification ─────────────┐
from models import (
    Module,
    Question,
    Respondent,
    Answer,
    Section,
    Survey,
    Template,
    Option,
    User,
)
from pydantic import BaseModel
from starlette.middleware.sessions import SessionMiddleware
from auth import router as auth_router, get_current_user, require_roles
from sondage_loader import load_sondage_complet
from services.export_csv import generate_csv_response
from services.visualisation_data import get_visualisation_context

load_dotenv()
# ┌─ Configuration ────────────────────────────────────────────────────────┐
# Les trois slugs de dashboard reconnus par l'application
VALID_ROLES = {"admin", "student", "program_manager"}


def role_to_dashboard_slug(role: str) -> str:
    """
    Convertit le rôle stocké en BDD en slug de route dashboard.

    "admin"              → "admin"
    "program_manager"              → "program_manager"
    "program_manager:MDE_P2027"    → "program_manager"
    "student" (ou autre) → "student"
    """
    if role.startswith("admin"):
        return "admin"
    elif role.startswith("program_manager"):
        return "program_manager"
    else:
        return "student"


def parse_rprm_formations(role: str) -> list[str]:
    """
    Extrait la liste des formations autorisées depuis une chaîne de rôle RP-RM.

    "program_manager:PROGRAM1;PROGRAM2" → ["PROGRAM1", "PROGRAM2"]
    "program_manager:PROGRAM1"            → ["PROGRAM1"]
    "program_manager"                       → []
    "admin"                       → []
    "admin:PROGRAM1;PROGRAM2" → ["PROGRAM1", "PROGRAM2"]
    "admin:PROGRAM1"            → ["PROGRAM1"]
    """
    if not role or not isinstance(role, str):
        return []
    role_upper = role.strip()
    if not (
        role_upper.startswith("program_manager:") or role_upper.startswith("admin:")
    ):
        return []
    after_colon = role_upper.split(":", 1)[1]
    return [f.strip() for f in after_colon.split(";") if f.strip()]


# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Configuration de la base de données ──────────────────────────────────┐
sqlite_file_name = "database/db_oceens.db"
sqlite_url = f"sqlite:///{sqlite_file_name}"

connect_args = {"check_same_thread": False, "timeout": 15}
engine = create_engine(sqlite_url, connect_args=connect_args)


def create_db_and_tables():
    SQLModel.metadata.create_all(engine)


def get_session():
    with Session(engine) as session:
        yield session


SessionDep = Annotated[Session, Depends(get_session)]
# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Modèles Pydantic pour les données entrantes ────────────────────────┐
class SurveyCreate(BaseModel):
    template_id: int
    campus: str
    program: str
    semester: str
    school_year: str
    user_id: Optional[int] = 1


class ProfesseurBase(BaseModel):
    id: int
    firstname: str
    name: str


class ModuleCreate(BaseModel):
    id: int
    name: str
    one_teacher_in_list: bool = False
    teachers: List[ProfesseurBase]


class UECreate(BaseModel):
    id: int
    name: str
    is_optional: bool
    modules: List[ModuleCreate]


class SurveyFullCreate(BaseModel):
    template_id: int
    campus: str
    program: str
    semester: str
    school_year: str
    ues: List[UECreate]


class AnswerItem(BaseModel):
    section_id: int
    question_id: int
    value: str
    module_id: Optional[int] = None
    teacher: Optional[str] = None


class SurveySubmission(BaseModel):
    answers: List[AnswerItem]


class RoleUpdate(BaseModel):
    role: str


import json

# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Fonctions utilitaires ──────────────────────────────────────────────┐
def parse_name(full_name: Optional[str], fallback_id: int) -> Dict[str, Optional[str]]:
    if not full_name:
        return {"id": fallback_id, "firstname": None, "name": None}
    parts = full_name.strip().split()
    if len(parts) == 1:
        return {"id": fallback_id, "firstname": parts[0], "name": ""}
    return {"id": fallback_id, "firstname": parts[0], "name": " ".join(parts[1:])}


def build_parametrage_data(
    session: Session, allowed_programs: list[str] | None = None
) -> Dict[str, object]:
    templates = session.exec(select(Template)).all()
    surveys = session.exec(select(Survey)).all()
    modules = session.exec(select(Module)).all()
    users = session.exec(select(User)).all()

    campus_names = []
    program_names = []
    semesters = []
    school_years = []
    formation_to_campus: Dict[str, str] = {}
    for survey in surveys:
        if survey.campus and survey.campus not in campus_names:
            campus_names.append(survey.campus)
        if survey.program and survey.program not in program_names:
            program_names.append(survey.program)
            formation_to_campus[survey.program] = survey.campus or ""
        if survey.semester and survey.semester not in semesters:
            semesters.append(survey.semester)
        if survey.school_year and survey.school_year not in school_years:
            school_years.append(survey.school_year)

    # Filtrer les filières si l'utilisateur RP-RM n'a accès qu'à certaines formations
    if allowed_programs is not None:
        program_names = [f for f in program_names if f in allowed_programs]
        # Ajouter les formations autorisées absentes des surveys existants
        for ap in allowed_programs:
            if ap not in program_names:
                program_names.append(ap)

    default_campuses = ["Paris-Cachan", "Montpellier", "Troyes", "St-Nazaire"]
    for dc in default_campuses:
        if dc not in campus_names:
            campus_names.append(dc)

    campus_list = [
        {"id": index + 1, "name": campus} for index, campus in enumerate(campus_names)
    ]
    campus_index = {campus["name"]: campus["id"] for campus in campus_list}
    programs = []
    for index, program in enumerate(program_names):
        programs.append(
            {
                "id": index + 1,
                "name": program,
                "campus_id": campus_index.get(
                    formation_to_campus.get(program, ""), None
                ),
            }
        )

    teachers = []
    teacher_index = 1
    teachers_seen = {}
    for module in modules:
        if not module.teacher:
            continue
        teachers_list_as_string = [
            p.strip() for p in module.teacher.split(",") if p.strip()
        ]
        for teacher_as_string in teachers_list_as_string:
            teacher = parse_name(teacher_as_string, teacher_index)
            if not teacher["firstname"] and not teacher["name"]:
                continue
            key = (teacher["firstname"].lower(), teacher["name"].lower())
            if key not in teachers_seen:
                teachers_seen[key] = teacher_index
                teacher["id"] = teacher_index
                teachers.append(teacher)
                teacher_index += 1

    for user in users:
        if user.role and "Enseignant" in user.role and user.mail:
            parsed = parse_name(
                user.mail.split("@")[0].replace(".", " "), teacher_index
            )
            parsed["name"] = parsed["name"] or ""
            parsed["firstname"] = parsed["firstname"] or ""
            key = (parsed["firstname"].lower(), parsed["name"].lower())
            if key and key not in teachers_seen:
                teachers_seen[key] = teacher_index
                parsed["id"] = teacher_index
                teachers.append(parsed)
                teacher_index += 1

    ues_by_program = {}
    if modules and programs:
        default_program_id = programs[0]["id"]
        for module in modules:
            program_id = default_program_id
            ue_name = module.ue or "Sans UE"
            ues_by_program.setdefault(program_id, [])
            ue_entry = next(
                (ue for ue in ues_by_program[program_id] if ue["name"] == ue_name), None
            )
            if ue_entry is None:
                ue_entry = {
                    "id": len(ues_by_program[program_id]) + 1,
                    "name": ue_name,
                    "is_optional": module.is_optional or False,
                    "_open": True,
                    "modules": [],
                }
                ues_by_program[program_id].append(ue_entry)
            teachers_list = []
            if module.teacher:
                teachers_list_as_string = [
                    p.strip() for p in module.teacher.split(",") if p.strip()
                ]
                for teacher_as_string in teachers_list_as_string:
                    parsed = parse_name(teacher_as_string, 0)
                    if parsed["firstname"] or parsed["name"]:
                        key = (parsed["firstname"].lower(), parsed["name"].lower())
                        if key not in teachers_seen:
                            teachers_seen[key] = teacher_index
                            parsed["id"] = teacher_index
                            teachers.append(parsed)
                            teacher_index += 1
                        teachers_list.append(
                            {
                                "id": teachers_seen[key],
                                "firstname": parsed["firstname"],
                                "name": parsed["name"],
                            }
                        )
            ue_entry["modules"].append(
                {
                    "id": int(module.module_id or 0),
                    "name": module.name or "Module",
                    "one_teacher_in_list": bool(module.one_teacher_in_list),
                    "teachers": teachers_list,
                }
            )

    template_dicts = [template.dict() for template in templates]

    return {
        "templates": template_dicts,
        "campus_list": campus_list,
        "programs": programs,
        "semesters": semesters,
        "school_years": school_years,
        "teachers_list": teachers,
        "ues_by_program": ues_by_program,
        "selected_template_id": template_dicts[0]["template_id"]
        if template_dicts
        else None,
        "selected_campus_id": campus_list[0]["id"] if campus_list else None,
        "selected_program_id": programs[0]["id"] if programs else None,
        "semester_year": semesters[0] if semesters else "",
        "selected_school_year": school_years[0] if school_years else "",
        "questions": [
            question.dict() for question in session.exec(select(Question)).all()
        ],
        "options": [option.dict() for option in session.exec(select(Option)).all()],
    }


# └────────────────────────────────────────────────────────────────────────┘


# ┌─ Gestion du cycle de vie (lifespan) ──────────────────────────────────┐
@asynccontextmanager
async def lifespan(app: FastAPI):
    print("Initialisation de la base de données...")
    create_db_and_tables()
    seed_all_if_necessary()
    yield
    print("Fermeture de la connexion...")


# └────────────────────────────────────────────────────────────────────────┘


def create_app():
    """
    Crée et configure l'application FastAPI fusionnée.
    """
    app = FastAPI(
        title="OceENS",
        description="Système de gestion et de connexion pour étudiants, professeurs et admins",
        lifespan=lifespan,
    )

    # SessionMiddleware (authentification)
    app.add_middleware(
        SessionMiddleware,
        secret_key=os.environ.get(
            "SECRET_KEY", "Y3mNqRjGQixkKjF9GXBCbOw2fHyC1wA3wqbJcQoIxt0="
        ),
        https_only=True,
        same_site="lax",
    )

    # Routeur d'authentification (login/logout/callback Azure Entra ID)
    app.include_router(auth_router)

    # Fichiers statiques et templates (montés une seule fois)
    app.mount("/static", StaticFiles(directory="static"), name="static")
    templates = Jinja2Templates(directory="templates")

    # ┌─ Route : Page d'accueil (version app.py conservée) ──────────────┐
    @app.get("/", response_class=HTMLResponse)
    async def index(request: Request):
        """
        Page d'accueil. Si l'utilisateur est déjà connecté avec un rôle
        valide, redirection vers son dashboard. Sinon, affichage du login.
        """
        user = get_current_user(request)
        if user and user.get("role"):
            slug = role_to_dashboard_slug(user["role"])
            return RedirectResponse(url=f"/dashboard/{slug}")
        return templates.TemplateResponse(request=request, name="index.html")

    # └────────────────────────────────────────────────────────────────┘

    dashboard_router = APIRouter(tags=["Dashboard"], prefix="/dashboard")

    api_router = APIRouter(tags=["API"], prefix="/api")

    # ┌─ Route : Paramétrage (accès restreint Admin + RP-RM) ──────────────┐
    @dashboard_router.get("/survey-create", response_class=HTMLResponse)
    def surveys_create(request: Request, session: SessionDep):
        # ── Sécurité : vérifier que l'utilisateur est Admin ou RP-RM ──
        user = require_roles(request, ["admin", "program_manager"])
        if user is None:
            # Utilisateur non connecté ou rôle insuffisant → redirection
            connected_user = get_current_user(request)
            if connected_user:
                slug = role_to_dashboard_slug(connected_user.get("role", ""))
                return RedirectResponse(url=f"/dashboard/{slug}")
            return RedirectResponse(url="/")

        # Déterminer les formations autorisées pour un RP-RM
        allowed_programs = None
        is_program_manager = False
        role = user.get("role", "") or ""

        if role.startswith("admin:"):
            allowed_programs = parse_rprm_formations(role)
            is_program_manager = True
        if role.startswith("program_manager:"):
            allowed_programs = parse_rprm_formations(role)
            is_program_manager = True
        elif role.startswith("program_manager"):
            is_program_manager = True

        data = build_parametrage_data(session, allowed_programs=allowed_programs)
        return templates.TemplateResponse(
            request=request,
            name="survey_create.html",
            context={
                "request": request,
                "templates": data["templates"],
                "campus_list": data["campus_list"],
                "programs": data["programs"],
                "semesters": data["semesters"],
                "school_years": data["school_years"],
                "teachers_list": data["teachers_list"],
                "ues_by_program": data["ues_by_program"],
                "selected_template_id": data["selected_template_id"],
                "selected_campus_id": data["selected_campus_id"],
                "selected_program_id": data["selected_program_id"],
                "semester_year": data["semester_year"],
                "selected_school_year": data["selected_school_year"],
                "is_program_manager": is_program_manager,
            },
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Données de paramétrage (accès restreint Admin + RP-RM) ────┐
    @api_router.get("/parametrage")
    def parametrage_api(request: Request, session: SessionDep):
        # ── Sécurité : vérifier que l'utilisateur est Admin ou RP-RM ──
        user = require_roles(request, ["admin", "program_manager"])
        if user is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin ou RP-RM requis."},
                status_code=403,
            )

        # Filtrer les filières pour les RP-RM
        allowed_programs = None
        role = user.get("role", "") or ""

        if ":" in role:  # RP-RM or Admin with formations
            allowed_programs = parse_rprm_formations(role)

        return JSONResponse(
            content=build_parametrage_data(session, allowed_programs=allowed_programs)
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Modules du survey de l'année précédente ─────────────────┐
    @api_router.get("/modules/previous")
    def modules_precedents_api(
        session: SessionDep,
        semester: str = "",
        program: str = "",
        school_year: str = "",
    ):
        """
        Retourne les modules du survey de l'année scolaire précédente
        pour le même semester et la même program.

        Exemple : school_year="2025-2026" → cherche "2024-2025".
        Si aucun survey précédent n'existe, renvoie un tableau vide.
        """
        if not semester or not program or not school_year:
            return JSONResponse(content={"ues": [], "teachers_list": []})

        # ── Calcul de l'année précédente ──────────────────────────────
        try:
            parts = school_year.split("-")
            if len(parts) == 2:
                year_start = int(parts[0]) - 1
                year_end = int(parts[1]) - 1
                previous_school_year = f"{year_start}-{year_end}"
            else:
                return JSONResponse(content={"ues": [], "teachers_list": []})
        except (ValueError, IndexError):
            return JSONResponse(content={"ues": [], "teachers_list": []})

        # ── Recherche du survey de l'année précédente ────────────────
        previous_survey = session.exec(
            select(Survey).where(
                Survey.school_year == previous_school_year,
                Survey.semester == semester,
                Survey.program == program,
            )
        ).first()

        if not previous_survey:
            return JSONResponse(content={"ues": [], "teachers_list": []})

        # ── Récupération des modules liés à ce survey ────────────────
        modules = session.exec(
            select(Module).where(
                Module.survey_id == previous_survey.survey_id,
                Module.template_id == previous_survey.template_id,
            )
        ).all()

        if not modules:
            return JSONResponse(content={"ues": [], "teachers_list": []})

        # ── Groupement par UE + extraction des professeurs ────────────
        ues_dict = {}
        teachers_seen = {}
        teachers = []
        teacher_index = 1

        for module in modules:
            ue_name = module.ue or "Sans UE"
            if ue_name not in ues_dict:
                ues_dict[ue_name] = {
                    "id": len(ues_dict) + 1,
                    "name": ue_name,
                    "is_optional": bool(module.is_optional),
                    "_open": True,
                    "modules": [],
                }

            teachers_list = []
            if module.teacher:
                teachers_list_as_string = [
                    p.strip() for p in module.teacher.split(",") if p.strip()
                ]
                for teacher_as_string in teachers_list_as_string:
                    parsed = parse_name(teacher_as_string, teacher_index)
                    if parsed["firstname"] or parsed["name"]:
                        key = (
                            (parsed["firstname"] or "").lower(),
                            (parsed["name"] or "").lower(),
                        )
                        if key not in teachers_seen:
                            teachers_seen[key] = teacher_index
                            parsed["id"] = teacher_index
                            teachers.append(parsed)
                            teacher_index += 1
                        teachers_list.append(
                            {
                                "id": teachers_seen[key],
                                "firstname": parsed["firstname"],
                                "name": parsed["name"],
                            }
                        )

            ues_dict[ue_name]["modules"].append(
                {
                    "id": int(module.module_id or 0),
                    "name": module.name or "Module",
                    "one_teacher_in_list": bool(module.one_teacher_in_list),
                    "teachers": teachers_list,
                }
            )

        return JSONResponse(
            content={
                "ues": list(ues_dict.values()),
                "teachersList": teachers_list,
                "previousSchoolYear": previous_school_year,
                "surveyId": previous_survey.survey_id,
            }
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Création d'un survey (accès restreint Admin + RP-RM) ────┐
    @api_router.post("/surveys")
    async def create_survey(
        request: Request,
        session: SessionDep,
        survey_data: str = Form(...),
        file: Optional[UploadFile] = File(None),
    ):
        """
        Crée un survey ET importe les étudiants en une seule transaction.
        Si l'import Excel échoue, le survey est annulé (ROLLBACK).
        """
        # ── Sécurité : vérifier que l'utilisateur est Admin ou RP-RM ──
        user = require_roles(request, ["admin", "program_manager"])
        if user is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin ou RP-RM requis."},
                status_code=403,
            )

        # ── Parse le JSON du survey envoyé en FormData ──
        try:
            survey_dict = json.loads(survey_data)
            survey = SurveyFullCreate(**survey_dict)
        except Exception as e:
            return JSONResponse(
                content={"error": f"Données du survey invalides : {str(e)}"},
                status_code=400,
            )

        # ── Sécurité : vérifier que la program est autorisée pour le RP-RM ──
        role = user.get("role", "")
        if ":" in role:  # RM-RP or Admin with formations
            allowed = parse_rprm_formations(role)
            if survey.program not in allowed:
                return JSONResponse(
                    content={
                        "error": f"Formation '{survey.program}' non autorisée pour votre rôle."
                    },
                    status_code=403,
                )

        # ── Pré-lecture du fichier Excel (avant la transaction) ──
        emails = []
        has_file = (
            file is not None
            and file.filename
            and file.filename.lower().endswith(".xlsx")
        )

        if file is not None and file.filename:
            if not file.filename.lower().endswith(".xlsx"):
                return JSONResponse(
                    content={
                        "error": "Format invalide. Seuls les fichiers .xlsx sont acceptés."
                    },
                    status_code=400,
                )

        if has_file:
            try:
                from openpyxl import load_workbook

                contents = await file.read()
                print(
                    f"[SONDAGE+IMPORT] Fichier reçu : {file.filename} ({len(contents)} octets)"
                )

                if len(contents) == 0:
                    return JSONResponse(
                        content={"error": "Le fichier Excel est vide."},
                        status_code=400,
                    )

                wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
                ws = wb.active
                print(f"[SONDAGE+IMPORT] Feuille active : {ws.title}")

                row_count = 0
                for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                    row_count += 1
                    cell_value = row[0]
                    if cell_value is None:
                        continue
                    cell_str = str(cell_value).strip()
                    if cell_str and "@" in cell_str:
                        emails.append(cell_str.lower())
                    elif cell_str and row_count <= 3:
                        print(
                            f"[SONDAGE+IMPORT] Ligne {row_count} ignorée (pas d'email) : '{cell_str}'"
                        )

                wb.close()
                print(
                    f"[SONDAGE+IMPORT] {row_count} ligne(s) lues, {len(emails)} email(s) valide(s)"
                )

                if not emails:
                    return JSONResponse(
                        content={
                            "error": f"Aucun email valide trouvé dans le fichier ({row_count} ligne(s) lue(s)). Vérifiez que les emails sont dans la première colonne."
                        },
                        status_code=400,
                    )

                # Dédupliquer
                emails = list(dict.fromkeys(emails))
                print(f"[SONDAGE+IMPORT] {len(emails)} email(s) unique(s) à traiter")

            except Exception as e:
                print(
                    f"[SONDAGE+IMPORT] Erreur lecture Excel : {type(e).__name__}: {e}"
                )
                return JSONResponse(
                    content={"error": f"Erreur de lecture du fichier Excel : {str(e)}"},
                    status_code=400,
                )

        # ── Transaction unique : Survey + Modules + Users + Respondent ──
        nb_crees = 0
        nb_existants = 0
        nb_repondre_inseres = 0

        try:
            with session.begin():
                with session.no_autoflush:
                    # ── Étape 1 : Créer le survey ──
                    existing_survey = session.exec(
                        select(Survey).where(Survey.template_id == survey.template_id)
                    ).all()
                    next_survey_id = (
                        max([s.survey_id for s in existing_survey] + [0]) + 1
                    )

                    survey_url = f"/api/surveys/{next_survey_id}"

                    new_survey = Survey(
                        template_id=survey.template_id,
                        survey_id=next_survey_id,
                        campus=survey.campus,
                        program=survey.program,
                        semester=survey.semester,
                        school_year=survey.school_year,
                        url=survey_url,
                        status=1,
                    )
                    session.add(new_survey)

                    # ── Étape 2 : Créer les modules ──
                    for ue in survey.ues:
                        for module_data in ue.modules:
                            prof_names = [
                                f"{p.firstname} {p.name}" for p in module_data.teachers
                            ]
                            enseignant_str = (
                                ", ".join(prof_names) if prof_names else None
                            )

                            new_module = Module(
                                name=module_data.name,
                                teacher=enseignant_str,
                                ue=ue.name,
                                is_optional=ue.is_optional,
                                one_teacher_in_list=module_data.one_teacher_in_list,
                                template_id=survey.template_id,
                                survey_id=next_survey_id,
                            )
                            session.add(new_module)

                    # ── Étape 3 : Importer les étudiants (si fichier fourni) ──
                    if emails:
                        email_to_user_id: Dict[str, int] = {}

                        existing_users = session.exec(select(User)).all()
                        existing_email_map = {
                            u.mail.lower(): u.user_id for u in existing_users if u.mail
                        }
                        print(
                            f"[SONDAGE+IMPORT] {len(existing_email_map)} user(s) existant(s) en BDD"
                        )

                        max_id = max([u.user_id for u in existing_users] + [0])

                        for email in emails:
                            if email in existing_email_map:
                                email_to_user_id[email] = existing_email_map[email]
                                nb_existants += 1
                            else:
                                max_id += 1
                                new_user = User(
                                    user_id=max_id,
                                    mail=email,
                                    role="student",
                                )
                                session.add(new_user)
                                email_to_user_id[email] = max_id
                                existing_email_map[email] = max_id
                                nb_crees += 1

                        print(
                            f"[SONDAGE+IMPORT] Users : {nb_crees} créé(s), {nb_existants} existant(s)"
                        )

                        user_ids = list(email_to_user_id.values())
                        if user_ids:
                            # On ne supprime PAS les anciennes affectations.
                            # Un étudiant peut être affecté à plusieurs sondages sur plusieurs années.
                            # On vérifie seulement si l'affectation existe déjà pour CE sondage.
                            existing_respondent_user_ids = session.exec(
                                select(Respondent.user_id).where(
                                    Respondent.template_id == survey.template_id,
                                    Respondent.survey_id == next_survey_id,
                                    Respondent.user_id.in_(user_ids),
                                )
                            ).all()

                            existing_respondent_user_ids = set(
                                existing_respondent_user_ids
                            )

                            # Insertion (INSERT) : ajouter uniquement les nouvelles affectations
                            for user_id in user_ids:
                                if user_id in existing_respondent_user_ids:
                                    continue  # Déjà affecté à ce survey, on ne fait rien
                                new_repondre = Respondent(
                                    template_id=survey.template_id,
                                    survey_id=next_survey_id,
                                    user_id=user_id,
                                    has_answered=False,
                                    submission_date=None,
                                )
                                session.add(new_repondre)
                                nb_repondre_inseres += 1

                        print(
                            f"[SONDAGE+IMPORT] Respondent : {nb_repondre_inseres} affectation(s) ajoutée(s)."
                        )

            # Si on arrive ici, le COMMIT a été fait par le context manager
            print(f"[SONDAGE+IMPORT] Transaction COMMIT réussie !")

        except Exception as e:
            print(f"[SONDAGE+IMPORT] ERREUR — ROLLBACK : {type(e).__name__}: {e}")
            import traceback

            traceback.print_exc()
            return JSONResponse(
                content={"error": f"Erreur lors de la création du survey : {str(e)}"},
                status_code=500,
            )

        result = {
            "message": "Survey créé avec succès",
            "survey_id": next_survey_id,
            "survey_url": survey_url,
        }
        if emails:
            result.update(
                {
                    "nb_emails_lus": len(emails),
                    "nb_users_crees": nb_crees,
                    "nb_users_existants": nb_existants,
                    "nb_repondre_inseres": nb_repondre_inseres,
                }
            )
        return result

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ Page questionnaire ─────────────────────────────────────────────┐
    @api_router.get("/surveys/{survey_id}", response_class=HTMLResponse)
    def questionnaire_page(request: Request, survey_id: int, session: SessionDep):
        survey = session.exec(
            select(Survey).where(
                Survey.survey_id == survey_id,
            )
        ).first()
        if not survey:
            return HTMLResponse(content="Survey introuvable.", status_code=404)

        sections = session.exec(
            select(Section)
            .where(Section.template_id == survey.template_id)
            .order_by(Section.order)
        ).all()
        questions = session.exec(
            select(Question).where(Question.template_id == survey.template_id)
        ).all()
        options = session.exec(
            select(Option).where(Option.template_id == survey.template_id)
        ).all()
        modules = session.exec(
            select(Module).where(Module.survey_id == survey_id)
        ).all()

        sections_data = []
        for sec in sections:
            sec_questions = [q for q in questions if q.section_id == sec.section_id]
            sec_questions.sort(key=lambda q: q.question_id)
            questions_data = []
            for q in sec_questions:
                q_options = [
                    o
                    for o in options
                    if o.section_id == sec.section_id and o.question_id == q.question_id
                ]
                q_options.sort(key=lambda o: o.option_id)
                questions_data.append(
                    {
                        "question_id": q.question_id,
                        "text": q.text,
                        "question_type": q.question_type,
                        "category": q.category,
                        "options": [
                            {"option_id": o.option_id, "text": o.text}
                            for o in q_options
                        ],
                    }
                )
            sections_data.append(
                {
                    "section_id": sec.section_id,
                    "name": sec.name,
                    "questions": questions_data,
                }
            )

        modules_data = []
        for mod in modules:
            teachers = []
            if mod.teacher:
                teachers = [p.strip() for p in mod.teacher.split(",") if p.strip()]
            modules_data.append(
                {
                    "module_id": mod.module_id,
                    "name": mod.name,
                    "ue": mod.ue,
                    "is_optional": bool(mod.is_optional),
                    "teachers": teachers,
                    "one_teacher_in_list": bool(mod.one_teacher_in_list),
                }
            )
        # Grouper les modules par UE pour la logique conditionnelle
        ues_data = {}
        for mod_data in modules_data:
            ue_name = mod_data["ue"] or "Sans UE"
            if ue_name not in ues_data:
                ues_data[ue_name] = {
                    "name": ue_name,
                    "is_optional": mod_data["is_optional"],
                    "modules": [],
                }
            ues_data[ue_name]["modules"].append(mod_data)
        ues_list = list(ues_data.values())

        return templates.TemplateResponse(
            request=request,
            name="survey.html",
            context={
                "request": request,
                "survey": {
                    "template_id": survey.template_id,
                    "survey_id": survey.survey_id,
                    "campus": survey.campus,
                    "program": survey.program,
                    "semester": survey.semester,
                    "school_year": survey.school_year,
                },
                "sections": sections_data,
                "modules": modules_data,
                "ues": ues_list,
            },
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Soumission des réponses du questionnaire ─────────────────┐
    @api_router.post("/surveys/{survey_id}")
    def submit_reponses(
        request: Request,
        survey_id: int,
        submission: SurveySubmission,
        session: SessionDep,
    ):
        # 1. Authentification : récupérer l'utilisateur connecté (Azure Entra ID)
        user = get_current_user(request)
        if not user:
            return JSONResponse(
                content={"error": "Authentification requise. Veuillez vous connecter."},
                status_code=401,
            )

        # 2. Résoudre l'user_id depuis l'email de l'utilisateur connecté
        db_user = session.exec(
            select(User).where(User.mail == user["email"].casefold())
        ).first()
        if not db_user:
            return JSONResponse(
                content={
                    "error": "Utilisateur "
                    + user["email"]
                    + "non trouvé dans la base de données."
                },
                status_code=403,
            )

        # 3. Vérifier que le survey existe
        survey = session.exec(
            select(Survey).where(
                Survey.survey_id == survey_id,
            )
        ).first()
        if not survey:
            return JSONResponse(
                content={"error": "Survey introuvable."}, status_code=404
            )

        # 4. Vérifier que cet élève est assigné à ce survey (table Respondent)
        #    Règle stricte : pas de INSERT, UPDATE uniquement
        respondent = session.exec(
            select(Respondent).where(
                Respondent.survey_id == survey_id,
                Respondent.user_id == db_user.user_id,
            )
        ).first()
        if not respondent:
            return JSONResponse(
                content={
                    "error": "Vous n'êtes pas autorisé ou assigné à répondre à ce survey."
                },
                status_code=403,
            )

        # 5. Vérifier que l'élève n'a pas déjà soumis ses réponses
        if (
            respondent.submission_date != None
        ):  # submission_date NOT NULL = has_answered
            return JSONResponse(
                content={"error": "Vous avez déjà soumis vos réponses pour ce survey."},
                status_code=409,
            )

        # 6. Enregistrement atomique : insertion des réponses + UPDATE Respondent
        #    On utilise begin_nested() (SAVEPOINT) car la session a déjà une
        #    transaction implicite ouverte par le générateur get_session().
        try:
            with session.begin_nested():
                # Calculer le prochain submission_id
                existing_reponses = session.exec(select(Answer)).all()
                existing_submission_ids = [
                    r.submission_id
                    for r in existing_reponses
                    if r.submission_id is not None
                ]
                submission_id = max(existing_submission_ids + [0]) + 1
                # TODO Mettre un lock / transaction pour éviter d'avoir le même submission_id par utilisateur

                # Insérer chaque réponse individuelle dans la table answers
                for rep in submission.answers:
                    new_reponse = Answer(
                        survey_id=survey_id,
                        section_id=rep.section_id,
                        module_id=rep.module_id,
                        teacher=rep.teacher,
                        question_id=rep.question_id,
                        submission_id=submission_id,
                        value=rep.value,
                    )
                    session.add(new_reponse)

                # UPDATE de la ligne Respondent : marquer comme répondu
                respondent.submission_date = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                session.add(respondent)

            # Commit de la transaction principale
            session.commit()

        except Exception as e:
            session.rollback()
            return JSONResponse(
                content={"error": f"Erreur lors de l'enregistrement : {str(e)}"},
                status_code=500,
            )

        return {
            "message": "Réponses enregistrées avec succès",
            # "user_id": db_user.user_id,
        }

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ Route : Dashboards par rôle ────────────────────────────────────┐

    @dashboard_router.get("/{role}", response_class=HTMLResponse)
    async def dashboard(request: Request, role: str, session: SessionDep):
        user = get_current_user(request)
        if not user:
            return RedirectResponse(url="/")
        if role not in VALID_ROLES:
            return RedirectResponse(url="/")

        user_slug = role_to_dashboard_slug(user.get("role", ""))
        if user_slug != "admin" and user_slug != role:  # Admin can see all the pages
            return RedirectResponse(url=f"/dashboard/{user_slug}")

        template_map = {
            "admin": "dashboard/admin.html",
            "student": "dashboard/student.html",
            "program_manager": "dashboard/program_manager.html",
        }

        programs = []
        full_role = user.get("role", "")
        if ":" in full_role:
            programs = [
                f.strip() for f in full_role.split(":", 1)[1].split(";") if f.strip()
            ]

        context = {"user": user, "programs": programs}

        if role == "admin":
            db_users = session.exec(select(User)).all()
            context["users"] = [
                {"user_id": u.user_id, "mail": u.mail, "role": u.role} for u in db_users
            ]

        if role in ("admin", "program_manager"):
            all_surveys = session.exec(select(Survey)).all()
            if role == "program_manager" and programs:
                surveys_filter = [s for s in all_surveys if s.program in programs]
            else:
                surveys_filter = list(all_surveys)

            surveys_list = []
            for s in surveys_filter:
                respondents_count = (
                    session.exec(
                        select(func.count(Respondent.user_id)).where(
                            Respondent.survey_id == s.survey_id,
                        )
                    ).first()
                    or 0
                )

                answers_count = (
                    session.exec(
                        select(func.count(Respondent.user_id)).where(
                            Respondent.survey_id == s.survey_id,
                            Respondent.submission_date
                            != None,  # submission_date NOT NULL = has_answered
                        )
                    ).first()
                    or 0
                )

                surveys_list.append(
                    {
                        "template_id": s.template_id,
                        "survey_id": s.survey_id,
                        "campus": s.campus,
                        "program": s.program,
                        "semester": s.semester,
                        "school_year": s.school_year,
                        "respondents_count": respondents_count,
                        "answers_count": answers_count,
                    }
                )
            context["surveys"] = surveys_list

        return templates.TemplateResponse(
            request=request,
            name=template_map[role],
            context=context,
        )

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Questionnaire(s) assigné à l'étudiant connecté ──────────────┐
    @api_router.get("/surveys/my/")
    def get_connected_user_surveys(request: Request, session: SessionDep):
        """
        Retourne tous les questionnaires assignés à l'étudiant connecté.
        Interroge la table Respondent pour récupérer les couples survey_id/template_id
        associés à l'étudiant, ainsi que son statut de réponse pour chaque sondage.
        """
        user = get_current_user(request)
        if not user:
            return JSONResponse(
                content={"error": "Authentification requise."},
                status_code=401,
            )

        # Résoudre l'user_id depuis l'email de l'utilisateur connecté
        db_user = session.exec(
            select(User).where(User.mail == user["email"].casefold())
        ).first()

        if not db_user:
            return JSONResponse(
                content={
                    "error": "Utilisateur "
                    + user["email"]
                    + " non trouvé en base de données."
                },
                status_code=404,
            )

        # Chercher toutes les entrées Respondent pour cet utilisateur
        # Un étudiant peut être affecté à plusieurs sondages, par exemple
        # sur plusieurs semestres ou plusieurs années scolaires.
        respondent_entries = session.exec(
            select(Respondent).where(Respondent.user_id == db_user.user_id)
        ).all()

        if not respondent_entries:
            return JSONResponse(
                content={
                    "surveys": [],
                    "message": "Aucun questionnaire assigné.",
                },
                status_code=200,
            )

        # Construire la liste des questionnaires assignés à l'étudiant
        surveys = []

        for entry in respondent_entries:
            # Récupérer les infos du survey pour le contexte d'affichage
            survey = session.exec(
                select(Survey).where(
                    Survey.survey_id == entry.survey_id,
                )
            ).first()

            # Sécurité : si la ligne Respondent pointe vers un survey inexistant,
            # on l'ignore pour éviter de casser la réponse API.
            if not survey:
                continue

            # submission_date NOT NULL = l'étudiant a déjà répondu
            has_answered = entry.submission_date is not None

            # status = 0 signifie que le sondage est fermé
            is_closed = survey.status == 0

            surveys.append(
                {
                    "template_id": survey.template_id,
                    "survey_id": survey.survey_id,
                    "has_answered": has_answered,
                    "url": f"/api/surveys/{survey.survey_id}",
                    "campus": survey.campus,
                    "program": survey.program,
                    "semester": survey.semester,
                    "school_year": survey.school_year,
                    "status": survey.status,
                    "is_closed": is_closed,
                    "can_answer": not has_answered and not is_closed,
                }
            )

        return {
            "surveys": surveys,
        }

    # └────────────────────────────────────────────────────────────────┘

    # ┌─ API : Gestion des rôles utilisateurs (accès restreint Admin) ────┐
    def _is_valid_role(role: str) -> bool:
        """Accepte 'admin' ou 'admin:program1,program2', 'student', 'program_manager' ou 'program_manager:program1,program2;...'"""
        if role in {"student"}:
            return True
        if role.startswith("admin"):
            return True
        if role.startswith("program_manager"):
            return True
        return False

    @api_router.get("/users")
    def get_users(request: Request, session: SessionDep):
        # ── Sécurité : seul un Admin peut lister tous les utilisateurs ──
        user = require_roles(request, ["admin"])
        if user is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin requis."},
                status_code=403,
            )
        users = session.exec(select(User)).all()
        return [{"user_id": u.user_id, "mail": u.mail, "role": u.role} for u in users]

    @api_router.put("/users/{user_id}/role")
    def update_user_role(
        request: Request, user_id: int, body: RoleUpdate, session: SessionDep
    ):
        # ── Sécurité : seul un Admin peut modifier les rôles ──
        admin = require_roles(request, ["admin"])
        if admin is None:
            return JSONResponse(
                content={"error": "Accès refusé. Rôle Admin requis."},
                status_code=403,
            )
        if not _is_valid_role(body.role):
            return JSONResponse(
                content={"detail": f"Rôle invalide : '{body.role}'"},
                status_code=422,
            )
        user = session.get(User, user_id)
        if not user:
            return JSONResponse(
                content={"detail": f"Utilisateur {user_id} introuvable"},
                status_code=404,
            )
        user.role = body.role
        session.add(user)
        session.commit()
        session.refresh(user)

        return {"user_id": user.user_id, "mail": user.mail, "role": user.role}

    # ┌─ Visualisation & Export CSV ──────────────────────────────────────┐
    def _check_sondage_access_and_status(
        session: Session,
        survey_id: int,
        role: str,
        allowed_programs: list[str],
    ):
        """Helper pour vérifier les accès et le statut de participation"""
        survey = session.exec(
            select(Survey).where(Survey.survey_id == survey_id)
        ).first()
        if not survey:
            return (
                None,
                {"error": "Survey introuvable.", "status_code": 404},
                None,
                None,
            )

        if role != "admin" and survey.program not in allowed_programs:
            return (
                None,
                {
                    "error": f"Formation '{survey.program}' non autorisée pour votre rôle.",
                    "status_code": 403,
                },
                None,
                None,
            )

        respondents_count = (
            session.exec(
                select(func.count(Respondent.user_id)).where(
                    Respondent.survey_id == survey_id,
                )
            ).first()
            or 0
        )
        answers_count = (
            session.exec(
                select(func.count(Respondent.user_id)).where(
                    Respondent.survey_id == survey_id,
                    Respondent.submission_date
                    != None,  # submission_date NOT NULL = has_answered
                )
            ).first()
            or 0
        )

        warning_msg = None
        if answers_count < respondents_count or survey.status == 1:
            warning_msg = f"Attention : Le sondage est toujours en cours. Seulement {answers_count} élève(s) ont répondu sur {respondents_count} inscrits."

        return survey, warning_msg, respondents_count, answers_count

    @api_router.get("/surveys/{survey_id}/export")
    def export_sondage_csv(request: Request, survey_id: int, session: SessionDep):
        user = require_roles(request, ["admin", "program_manager"])
        if user is None:
            return JSONResponse(content={"error": "Accès refusé."}, status_code=403)

        role = user.get("role", "") or ""
        allowed_programs = parse_rprm_formations(role) if ":" in role else []
        admin_role = "admin" if role == "admin" else "program_manager"

        survey, error_or_warning, _, _ = _check_sondage_access_and_status(
            session, survey_id, admin_role, allowed_programs
        )
        if not survey:
            return JSONResponse(
                content={"error": error_or_warning["error"]},
                status_code=error_or_warning["status_code"],
            )

        # Utilisation de la BDD locale pour le loader sqlite3 natif
        survey_obj = load_sondage_complet("database/db_oceens.db", survey_id)

        resp = generate_csv_response(survey_obj)
        if isinstance(error_or_warning, str):
            resp.headers["X-Warning"] = "Survey en cours - donnees partielles"

        return resp

    @api_router.get("/surveys/{survey_id}/visualisation", response_class=HTMLResponse)
    def visualisation_page(request: Request, survey_id: int, session: SessionDep):
        user = require_roles(request, ["admin", "program_manager"])
        if user is None:
            return RedirectResponse(url="/")

        role = user.get("role", "") or ""
        allowed_programs = parse_rprm_formations(role) if ":" in role else []
        admin_role = "admin" if role == "admin" else "program_manager"

        survey, error_or_warning, respondents_count, answers_count = (
            _check_sondage_access_and_status(
                session, survey_id, admin_role, allowed_programs
            )
        )
        if not survey:
            return HTMLResponse(
                content=f"<h1>Erreur</h1><p>{error_or_warning['error']}</p>",
                status_code=error_or_warning["status_code"],
            )

        # Utilisation de la BDD locale pour le loader sqlite3 natif
        survey_obj = load_sondage_complet("database/db_oceens.db", survey_id)

        viz_context = get_visualisation_context(survey_obj)

        context = {
            "user": user,
            "survey": survey,
            "respondents_count": respondents_count,
            "answers_count": answers_count,
            "warning_msg": error_or_warning
            if isinstance(error_or_warning, str)
            else None,
            "viz_data": viz_context,
        }

        return templates.TemplateResponse(
            request=request, name="visualisation.html", context=context
        )

    # └───────────────────────────────────────────────────────────────────┘

    app.include_router(api_router)
    app.include_router(dashboard_router)

    return app


# ┌─ Instance applicative globale ───────────────────────────────────────┐
app = create_app()
# └──────────────────────────────────────────────────────────────────────┘


if __name__ == "__main__":
    uvicorn.run(
        "app:app",
        host="0.0.0.0",
        port=8000,
        reload=False,
    )
